"""
世纪互联太仓设备管理平台 - Flask后端
"""
import os
import re
import json
import pickle
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, send_file, session
from sqlalchemy import and_, or_, inspect, text
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook

from dcim_app.enums import (
    ASSET_STATUS_ENUM,
    GENERAL_ASSET_STATUS_ENUM,
    IT_RACK_MOUNT_STATUS_ENUM,
    PDU_ASSET_STATUS_ENUM,
)
from dcim_app.extensions import db, login_manager
from dcim_app.models import User, DeviceCategory, Device, AssetEvent
from dcim_app.permissions import (
    allowed_professions_for_current_user,
    can_access_maintenance,
    can_export,
    is_admin,
)
from dcim_app.utils.schema import ensure_device_schema_sqlite
from dcim_app.utils.dates import parse_date, parse_datetime
from dcim_app.utils.io import read_excel_rows as _read_excel_rows, read_pkl_rows as _read_pkl_rows
from dcim_app.utils.ids import (
    is_blank as _is_blank,
    to_str as _to_str,
    normalize_unique_id as _normalize_unique_id,
    unique_id_from_row as _unique_id_from_row,
)
from dcim_app.services.derived_fields import (
    parse_unique_id_meta as _parse_unique_id_meta,
    model_to_classification as _model_to_classification,
    backfill_uid_derived_fields,
)
from dcim_app.services.dashboard_stats import compute_it_cabinet_dashboard
from dcim_app.services.reimport_power import reimport_power_pkls
from dcim_app.services.warehouse_overrides import (
    load_ac_pdu_override_from_pkl as _load_ac_pdu_override_from_pkl,
    load_it_cabinet_override_from_pkl as _load_it_cabinet_override_from_pkl,
)
from dcim_app.services.warehouse_sync import load_devices_from_warehouse
from dcim_app.services.device_convert_params import lookup_convert_params

# 配置
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
ALLOWED_EXTENSIONS = {'xlsx'}
DEVICE_FILES = {
    '变压器': os.path.join(BASE_DIR, '..', 'R楼-变压器.xlsx'),
    'UPS': os.path.join(BASE_DIR, '..', 'R楼-UPS.xlsx'),
    '柴发': os.path.join(BASE_DIR, '..', 'R楼-柴发.xlsx')
}

STAFF_FILE = os.path.join(BASE_DIR, '..', '太仓运维人员名单.xlsx')

# 资产状态/上下架枚举已迁移至 dcim_app.enums

# 设备台账“仓库源”目录（优先用环境变量覆盖）
# 你提供的路径可通过设置环境变量 ASSET_WAREHOUSE_DIR 来指定。
WAREHOUSE_DIR = os.environ.get('ASSET_WAREHOUSE_DIR') or os.path.join(BASE_DIR, 'device_database')
# Excel→pkl 转换目录（含「参数」列的台账，用于设备详情补充展示）
DEVICE_CONVERT_DIR = os.environ.get('DEVICE_CONVERT_DIR') or os.path.join(
    BASE_DIR, 'device_convert'
)

app = Flask(__name__)
app.config['SECRET_KEY'] = 'DC-Asset-Manager-Secret-2024'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///asset_manager.db'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max

# 确保上传目录存在
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

db.init_app(app)
login_manager.init_app(app)
login_manager.login_view = 'login'


@app.before_request
def _ensure_device_schema_once():
    if app.config.get('_device_schema_v1'):
        return
    try:
        ensure_device_schema_sqlite()
    except Exception:
        pass
    app.config['_device_schema_v1'] = True

# ============== 数据库模型 ==============
# 已迁移至 dcim_app.models；此文件仅保留业务逻辑与路由

# ============== 辅助函数 ==============

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/export_tools')
@login_required
def export_tools():
    if not can_export():
        return jsonify({'success': False, 'message': '无导出权限'}), 403
    categories = DeviceCategory.query.order_by(DeviceCategory.display_name.asc().nullslast(), DeviceCategory.name.asc()).all()
    return render_template('export_tools.html', categories=categories)


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not is_admin():
            return jsonify({'error': '需要管理员权限'}), 403
        return f(*args, **kwargs)
    return decorated_function

 # parse_date / parse_datetime 已迁移至 dcim_app.utils.dates

 # _is_blank / _to_str / _normalize_unique_id / _unique_id_from_row 已迁移至 dcim_app.utils.ids


def _coerce_rack_mount_status(val) -> str | None:
    s = _to_str(val)
    if not s:
        return None
    if s in IT_RACK_MOUNT_STATUS_ENUM:
        return s
    for k in IT_RACK_MOUNT_STATUS_ENUM:
        if s.startswith(k):
            return k
    # 历史动环/旧版枚举兼容（统一落入 已上架 / 停用 / 未上架）
    legacy = {
        '已下架': '未上架',
        '待上架': '未上架',
        '预留': '未上架',
        '未标注': '未上架',
        '停用中': '停用',
        '使用中': '已上架',  # 上架状态列常见与资产状态混用
    }
    if s in legacy:
        return legacy[s]
    for old, new in legacy.items():
        if old and s.startswith(old):
            return new
    return None


RACK_MOUNT_ROW_KEYS = (
    'IT机柜上下架状态',
    '上下架状态',
    '机柜上下架状态',
    '上架状态',
    '机柜状态',
)


def _coerce_rack_mount_from_row(row: dict) -> str | None:
    """从行字典中取第一个可规范化的上下架状态（多列并存时按优先级）。"""
    if not isinstance(row, dict):
        return None
    for rk_key in RACK_MOUNT_ROW_KEYS:
        if rk_key in row and not _is_blank(row.get(rk_key)):
            coerced = _coerce_rack_mount_status(row.get(rk_key))
            if coerced:
                return coerced
    return None


 # reimport_power_pkls 及其 pkl 列名兼容逻辑已迁移至 dcim_app.services.reimport_power


def _rack_status_merge_rank(st: str | None) -> int:
    """合并多行时：停用 > 已上架 > 未上架，避免仅因行顺序导致 停用 被末行空值覆盖。"""
    if not st:
        return 0
    return {'未上架': 1, '已上架': 2, '停用': 3}.get(st, 0)


def _merge_it_cabinet_pkl_rows_by_unique_id(rows: list) -> list[dict]:
    """
    R楼-IT机柜 源常含同一唯一ID多行（导出/合并产生）。合并规则：
    - 普通字段：后者非空覆盖前者；
    - 上下架：取优先级最高的一行（保证 Excel 中「停用」不会因后续空行丢失）。
    """
    by_uid: dict[str, list[dict]] = defaultdict(list)
    no_uid: list[dict] = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        r = _normalize_pkl_row_keys(r)
        uid = _unique_id_from_row(r)
        if uid:
            by_uid[uid].append(r)
        else:
            no_uid.append(r)
    merged: list[dict] = []
    for group in by_uid.values():
        combined: dict = {}
        best_rank = 0
        best_rack: str | None = None
        for r in group:
            st = _coerce_rack_mount_from_row(r)
            rk = _rack_status_merge_rank(st)
            if rk > best_rank:
                best_rank = rk
                best_rack = st
            for k, v in r.items():
                if not _is_blank(v):
                    combined[k] = v
        if best_rack:
            combined['IT机柜上下架状态'] = best_rack
        merged.append(combined)
    return merged + no_uid


def _pkl_asset_status_in_use(st: str | None) -> bool:
    s = _to_str(st)
    if not s:
        return False
    return str(s).strip().startswith("使用中")


def _pkl_it_cabinet_row_in_use(row: dict) -> bool:
    if _pkl_asset_status_in_use(row.get("资产状态")):
        return True
    rk = _coerce_rack_mount_from_row(row)
    return rk == "已上架"


def _dedupe_pkl_rows_by_unique_id(rows: list) -> list[dict]:
    by_uid: dict[str, dict] = {}
    no_uid: list[dict] = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        uid = _unique_id_from_row(r)
        if uid:
            by_uid[uid] = r
        else:
            no_uid.append(r)
    return list(by_uid.values()) + no_uid


def ensure_it_cabinet_category() -> DeviceCategory:
    c = DeviceCategory.query.filter_by(name="IT机柜").first()
    if c:
        return c
    c = DeviceCategory(
        name="IT机柜",
        display_name="IT机柜",
        description="R楼-IT机柜.pkl 仓库同步",
    )
    db.session.add(c)
    db.session.commit()
    return c


def hydrate_it_cabinets_from_pkl(warehouse_dir: str | None = None) -> dict:
    """
    若库中「IT机柜」分类下尚无记录，则从 R楼-IT机柜.pkl 一次性灌入（与仓库同步逻辑一致）。
    解决仅存在 pkl 覆盖数、但明细/API 为空的问题。
    """
    warehouse_dir = warehouse_dir or WAREHOUSE_DIR
    p = os.path.join(warehouse_dir, "R楼-IT机柜.pkl")
    if not os.path.exists(p):
        return {"ok": False, "reason": "no_pkl"}
    cat = ensure_it_cabinet_category()
    if Device.query.filter_by(category_id=cat.id).count() > 0:
        return {"ok": True, "skipped": True, "reason": "db_already_has_rows"}
    try:
        rows = _read_pkl_rows(p)
        rows = _merge_it_cabinet_pkl_rows_by_unique_id(rows)
    except Exception as e:
        return {"ok": False, "error": str(e)}
    inserted = 0
    updated = 0
    for row in rows:
        if not isinstance(row, dict):
            continue
        category = cat
        unique_id = _unique_id_from_row(row)
        if unique_id:
            existing = Device.query.filter_by(category_id=category.id, unique_id=unique_id).first()
            if not existing:
                existing = Device.query.filter_by(unique_id=unique_id).first()
                if existing:
                    existing.category_id = category.id
            if existing:
                update_device_from_row(existing, row)
                updated += 1
            else:
                device = Device(category_id=category.id)
                update_device_from_row(device, row)
                db.session.add(device)
                inserted += 1
        else:
            device = Device(category_id=category.id)
            update_device_from_row(device, row)
            db.session.add(device)
            inserted += 1
    db.session.commit()
    normalize_it_rack_mount_status_for_it_cabinets()
    print(f"已从 R楼-IT机柜.pkl 灌入 IT机柜：新增 {inserted} 更新 {updated}")
    return {"ok": True, "inserted": inserted, "updated": updated, "rows": len(rows)}


def maybe_hydrate_it_cabinets_from_pkl():
    """在仪表盘 / API 访问时自动补数（仅当 IT机柜 为空）。"""
    try:
        return hydrate_it_cabinets_from_pkl()
    except Exception as e:
        print(f"IT机柜 pkl 灌入跳过: {e}")
        return {"ok": False, "error": str(e)}


def normalize_it_rack_mount_status_for_it_cabinets() -> int:
    """
    将「IT机柜」台账的 it_rack_mount_status 规范为 IT_RACK_MOUNT_STATUS_ENUM。
    空值、未标注及无法识别的历史值 → 未上架。
    """
    cat = DeviceCategory.query.filter_by(name="IT机柜").first()
    if not cat:
        return 0
    changed = 0
    for d in Device.query.filter_by(category_id=cat.id).all():
        raw = (d.it_rack_mount_status or "").strip()
        if not raw or raw == "未标注":
            if d.it_rack_mount_status != "未上架":
                d.it_rack_mount_status = "未上架"
                changed += 1
            continue
        if raw in IT_RACK_MOUNT_STATUS_ENUM:
            continue
        coerced = _coerce_rack_mount_status(raw)
        nv = coerced if coerced else "未上架"
        if d.it_rack_mount_status != nv:
            d.it_rack_mount_status = nv
            changed += 1
    if changed:
        db.session.commit()
    return changed


def _db_category_active_count(cat: DeviceCategory) -> int:
    """使用中：资产状态以「使用中」开头（含 使用中-已上电…）；IT机柜另计「已上架」。"""
    q = Device.query.filter_by(category_id=cat.id)
    in_use_asset = and_(
        Device.asset_status.isnot(None),
        Device.asset_status.like("使用中%"),
    )
    if cat.name == "IT机柜":
        return q.filter(or_(in_use_asset, Device.it_rack_mount_status == "已上架")).count()
    return q.filter(in_use_asset).count()


def ensure_device_schema_sqlite():
    """SQLite 增量加列（旧库无 it_rack_mount_status 时）。"""
    try:
        url = str(db.engine.url)
    except Exception:
        return
    if 'sqlite' not in url:
        return
    try:
        insp = inspect(db.engine)
        cols = {c['name'] for c in insp.get_columns('device')}
        to_add = []
        if 'it_rack_mount_status' not in cols:
            to_add.append(('it_rack_mount_status', 'VARCHAR(50)'))
        if 'object_kind' not in cols:
            to_add.append(('object_kind', 'VARCHAR(20)'))
        if 'profession' not in cols:
            to_add.append(('profession', 'VARCHAR(20)'))
        if 'model_classification' not in cols:
            to_add.append(('model_classification', 'VARCHAR(100)'))
        if not to_add:
            return
        with db.engine.begin() as conn:
            for col, typ in to_add:
                conn.execute(text(f'ALTER TABLE device ADD COLUMN {col} {typ}'))
    except Exception as e:
        print(f"schema migrate skipped: {e}")


 # _parse_unique_id_meta / _model_to_classification / backfill_uid_derived_fields 已迁移至 dcim_app.services.derived_fields

 # compute_it_cabinet_dashboard 已迁移至 dcim_app.services.dashboard_stats


 # _load_ac_pdu_override_from_pkl / _load_it_cabinet_override_from_pkl 已迁移至 dcim_app.services.warehouse_overrides


 # _read_excel_rows / _read_pkl_rows 已迁移至 dcim_app.utils.io

def _guess_category_from_filename(file_path: str) -> str:
    """
    Derive category name from xlsx filename.
    Examples: "R楼-UPS.xlsx" -> "UPS", "R楼-柴发.xlsx" -> "柴发"
    """
    stem = os.path.splitext(os.path.basename(file_path))[0].strip()
    for sep in ('-', '－', '—', '_', ' '):
        if sep in stem:
            parts = [p.strip() for p in stem.split(sep) if p and p.strip()]
            if parts:
                return parts[-1]
    return stem or '其他设备'


def _warehouse_stem_normalized(file_path: str) -> str:
    stem = os.path.splitext(os.path.basename(file_path))[0].strip()
    return stem.lower().replace(' ', '')


def is_r_it_cabinet_ac_pdu_warehouse_file(file_path: str) -> bool:
    """仓库源 R楼-IT机柜交流PDU.pkl / xlsx 等：台账统一归入「交流PDU」，而非 IT机柜交流PDU / 交流列头柜。"""
    n = _warehouse_stem_normalized(file_path)
    return 'r楼' in n and 'it机柜交流pdu' in n


def is_r_it_cabinet_only_warehouse_file(file_path: str) -> bool:
    """R楼-IT机柜.pkl：IT 机柜台账（排除 IT机柜交流PDU 文件名）。"""
    n = _warehouse_stem_normalized(file_path)
    if 'r楼' not in n or 'it机柜' not in n:
        return False
    if 'it机柜交流pdu' in n:
        return False
    return True


 # 仓库同步与分类修正已迁移至 dcim_app.services.warehouse_sync

def _staff_extract_name_empid(row: dict):
    name = _to_str(row.get('姓名') or row.get('name') or row.get('Name') or row.get('人员姓名') or row.get('人员'))
    emp_id_val = row.get('工号') or row.get('employee_id') or row.get('EmployeeId') or row.get('工单号') or row.get('员工工号')
    emp_id = None
    if not _is_blank(emp_id_val):
        try:
            emp_id = str(int(emp_id_val))
        except Exception:
            emp_id = _to_str(emp_id_val)
    return name, emp_id

def sync_users_from_staff_rows(rows: list[dict]):
    """
    Upsert staff list into DB users:
    - username = 姓名
    - password = V + 工号 (when工号存在)
    - role forced to user, active true
    """
    inserted = 0
    updated = 0
    skipped = 0
    for row in rows:
        name, emp_id = _staff_extract_name_empid(row)
        if not name:
            skipped += 1
            continue

        user = User.query.filter_by(username=name).first()
        created = False
        if not user:
            user = User(username=name, role='user', display_name=name)
            db.session.add(user)
            created = True

        user.display_name = name
        user.employee_id = emp_id
        user.role = 'user'
        user.is_active = True
        if emp_id:
            user.set_password(f'V{emp_id}')

        if created:
            inserted += 1
        else:
            updated += 1

    db.session.commit()
    return {'inserted': inserted, 'updated': updated, 'skipped': skipped, 'total': len(rows)}

def write_staff_excel(rows: list[dict], out_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = 'staff'
    ws.append(['姓名', '工号'])
    seen = set()
    for row in rows:
        name, emp_id = _staff_extract_name_empid(row)
        if not name or not emp_id:
            continue
        key = (name, emp_id)
        if key in seen:
            continue
        seen.add(key)
        ws.append([name, emp_id])
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)

def serialize_device(d: Device):
    return {
        'id': d.id,
        'scope_id': d.scope_id,
        'unique_id': d.unique_id,
        'instance_name': d.instance_name,
        'asset_status': d.asset_status,
        'location': d.location,
        'brand': d.brand,
        'model': d.model,
        'warranty_years': d.warranty_years,
        'lifecycle': d.lifecycle,
        'commission_date': d.commission_date.isoformat() if d.commission_date else None,
        'warranty_start': d.warranty_start.isoformat() if d.warranty_start else None,
        'warranty_end': d.warranty_end.isoformat() if d.warranty_end else None,
        'is_expired': d.is_expired,
        'energy_type': d.energy_type,
        'threshold_range': d.threshold_range,
        'device_type': d.device_type,
        'rated_capacity': d.rated_capacity,
        'qr_code': d.qr_code,
        'it_rack_mount_status': d.it_rack_mount_status,
        'object_kind': d.object_kind,
        'profession': d.profession,
        'model_classification': d.model_classification,
        'category': d.category.name if d.category else None,
        'category_id': d.category.id if d.category else None,
        'updated_at': d.updated_at.isoformat() if d.updated_at else None,
    }

def serialize_event(e: AssetEvent):
    payload = None
    if e.payload_json:
        try:
            payload = json.loads(e.payload_json)
        except Exception:
            payload = None
    return {
        'id': e.id,
        'device_id': e.device_id,
        'event_type': e.event_type,
        'title': e.title,
        'status': e.status,
        'occurred_at': e.occurred_at.isoformat() if e.occurred_at else None,
        'created_by': e.created_by,
        'note': e.note,
        'payload': payload,
        'created_at': e.created_at.isoformat() if e.created_at else None,
    }

def load_devices_from_excel():
    """从Excel文件加载设备数据"""
    for category_name, file_path in DEVICE_FILES.items():
        if not os.path.exists(file_path):
            continue
        
        # 检查分类是否存在
        category = DeviceCategory.query.filter_by(name=category_name).first()
        if not category:
            category = DeviceCategory(
                name=category_name,
                display_name=category_name,
                description=f'{category_name}设备台账'
            )
            db.session.add(category)
            db.session.commit()
        
        try:
            rows = _read_excel_rows(file_path)
            for row in rows:
                # 检查设备是否已存在
                unique_id = _unique_id_from_row(row)
                existing = Device.query.filter_by(
                    category_id=category.id,
                    unique_id=unique_id
                ).first()
                
                if existing:
                    # 更新现有设备
                    update_device_from_row(existing, row)
                else:
                    # 创建新设备
                    device = Device(category_id=category.id)
                    update_device_from_row(device, row)
                    db.session.add(device)
            
            db.session.commit()
            print(f"已加载 {category_name}: {len(rows)} 条记录")
        except Exception as e:
            print(f"加载 {category_name} 失败: {e}")

 # load_devices_from_warehouse 已迁移至 dcim_app.services.warehouse_sync

def update_device_from_row(device, row):
    """从Excel行更新设备数据"""
    device.scope_id = _to_str(row.get('范围域号'))
    device.unique_id = _unique_id_from_row(row)
    device.instance_name = _to_str(row.get('实例名称'))
    device.asset_status = _to_str(row.get('资产状态'))
    device.location = _to_str(row.get('位置信息'))
    device.brand = _to_str(row.get('品牌'))
    device.model = _to_str(row.get('型号'))
    device.warranty_years = _to_str(row.get('质保年限'))
    device.lifecycle = _to_str(row.get('生命周期'))
    device.commission_date = parse_date(row.get('投产日期'))
    device.warranty_start = parse_date(row.get('质保开始时间'))
    device.warranty_end = parse_date(row.get('质保结束时间'))
    device.energy_type = _to_str(row.get('能耗类型'))
    device.threshold_range = _to_str(row.get('阈值范围'))
    device.device_type = _to_str(row.get('动环设备类型'))
    device.rated_capacity = _to_str(row.get('额定电流容量'))
    device.qr_code = _to_str(row.get('二维码'))

    # 唯一ID派生字段
    meta = _parse_unique_id_meta(device.unique_id)
    device.object_kind = meta["object_kind"]
    device.profession = meta["profession"]
    device.model_classification = _model_to_classification(device.model)

    coerced_rack = _coerce_rack_mount_from_row(row)
    if coerced_rack:
        device.it_rack_mount_status = coerced_rack
    
    # 判断是否过保
    if device.warranty_end:
        device.is_expired = datetime.now().date() > device.warranty_end
    else:
        device.is_expired = False

def init_users():
    """初始化用户数据"""
    # 管理员账户
    if not User.query.filter_by(username='admin').first():
        admin = User(
            username='admin',
            display_name='系统管理员',
            role='admin'
        )
        admin.set_password('LJJ1993')
        db.session.add(admin)
    
    # 从Excel加载运维人员
    if os.path.exists(STAFF_FILE):
        try:
            rows = _read_excel_rows(STAFF_FILE)
            stats = sync_users_from_staff_rows(rows)
            print(f"用户初始化完成：新增{stats['inserted']} 更新{stats['updated']} 跳过{stats['skipped']}")
        except Exception as e:
            print(f"加载用户失败: {e}")
            db.session.rollback()

# ============== 路由 - 认证 ==============

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        data = request.get_json() if request.is_json else request.form
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        if not username or not password:
            return jsonify({'success': False, 'message': '请输入用户名和密码'})
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password) and user.is_active:
            login_user(user)
            next_page = request.args.get('next')
            return jsonify({'success': True, 'message': '登录成功', 'next': next_page or url_for('dashboard')})
        else:
            return jsonify({'success': False, 'message': '用户名或密码错误'})
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# ============== 路由 - 仪表盘 ==============

@app.route('/dashboard')
@login_required
def dashboard():
    ensure_it_cabinet_category()
    maybe_hydrate_it_cabinets_from_pkl()
    try:
        backfill_uid_derived_fields(limit=3000)
    except Exception:
        pass

    ac_pdu_override = _load_ac_pdu_override_from_pkl(
        warehouse_dir=WAREHOUSE_DIR,
        read_pkl_rows=_read_pkl_rows,
    )
    it_cab_override = _load_it_cabinet_override_from_pkl(
        warehouse_dir=WAREHOUSE_DIR,
        read_pkl_rows=_read_pkl_rows,
        merge_it_cabinet_rows=_merge_it_cabinet_pkl_rows_by_unique_id,
    )

    # 设备统计
    db_total = Device.query.count()
    db_active = Device.query.filter(
        Device.asset_status.isnot(None),
        Device.asset_status.like("使用中%"),
    ).count()
    db_expired = Device.query.filter_by(is_expired=True).count()
    db_categories = DeviceCategory.query.count()

    stats_total = db_total
    stats_active = db_active

    if ac_pdu_override:
        cat_ac_pdu = DeviceCategory.query.filter_by(name='交流PDU').first()
        if cat_ac_pdu:
            db_ac_total = Device.query.filter_by(category_id=cat_ac_pdu.id).count()
            db_ac_active = _db_category_active_count(cat_ac_pdu)
            stats_total = stats_total - db_ac_total + ac_pdu_override["total"]
            stats_active = stats_active - db_ac_active + ac_pdu_override["active"]

    if it_cab_override:
        cat_it = DeviceCategory.query.filter_by(name='IT机柜').first()
        if cat_it:
            db_it_total = Device.query.filter_by(category_id=cat_it.id).count()
            db_it_active = _db_category_active_count(cat_it)
            stats_total = stats_total - db_it_total + it_cab_override["total"]
            stats_active = stats_active - db_it_active + it_cab_override["active"]

    it_cat = DeviceCategory.query.filter_by(name='IT机柜').first()
    it_cabinet_total = Device.query.filter_by(category_id=it_cat.id).count() if it_cat else 0
    it_cabinet_active = _db_category_active_count(it_cat) if it_cat else 0
    if it_cab_override:
        it_cabinet_total = it_cab_override["total"]
        it_cabinet_active = it_cab_override["active"]

    stats = {
        'total': stats_total,
        'active': stats_active,
        'expired': db_expired,
        'categories': db_categories,
        'it_cabinet_total': it_cabinet_total,
        'it_cabinet_active': it_cabinet_active,
    }
    
    # 分类统计（IT机柜 保证有分类记录以便在「设备分类」网格中展示）
    ensure_it_cabinet_category()
    category_stats = []
    for cat in DeviceCategory.query.order_by(DeviceCategory.name.asc()).all():
        cat_devices = Device.query.filter_by(category_id=cat.id)
        total = cat_devices.count()
        active = _db_category_active_count(cat)
        expired = cat_devices.filter_by(is_expired=True).count()

        if ac_pdu_override and cat.name == '交流PDU':
            total = ac_pdu_override["total"]
            active = ac_pdu_override["active"]

        if it_cab_override and cat.name == 'IT机柜':
            total = it_cab_override["total"]
            active = it_cab_override["active"]

        category_stats.append({
            'id': cat.id,
            'name': cat.name,
            'display_name': cat.display_name,
            'total': total,
            'active': active,
            'expired': expired
        })

    category_stats.sort(
        key=lambda x: (0 if x["name"] == "IT机柜" else 1, x["display_name"] or x["name"] or "")
    )
    
    return render_template(
        'dashboard.html',
        stats=stats,
        category_stats=category_stats,
    )


@app.route('/it_cabinet_rack')
@login_required
def it_cabinet_rack():
    """IT 机柜上下架：饼图 + 明细；数据来自分类「IT机柜」（仓库源 R楼-IT机柜.pkl 同步）。"""
    ensure_it_cabinet_category()
    maybe_hydrate_it_cabinets_from_pkl()
    it_cat = DeviceCategory.query.filter_by(name='IT机柜').first()
    return render_template(
        'cabinet_rack.html',
        it_category_id=it_cat.id if it_cat else None,
        rack_mount_enum=IT_RACK_MOUNT_STATUS_ENUM,
        warehouse_pkl_hint=os.path.join(WAREHOUSE_DIR, 'R楼-IT机柜.pkl'),
    )


@app.route('/api/it_cabinets', methods=['GET'])
@login_required
def api_it_cabinets():
    ensure_it_cabinet_category()
    maybe_hydrate_it_cabinets_from_pkl()
    cat = DeviceCategory.query.filter_by(name='IT机柜').first()
    if not cat:
        return jsonify({'success': True, 'data': [], 'total': 0, 'chart': []})
    rows = Device.query.filter_by(category_id=cat.id).order_by(Device.updated_at.desc()).limit(5000).all()
    data = [serialize_device(d) for d in rows]
    bucket = Counter()
    for d in rows:
        raw = (d.it_rack_mount_status or '').strip()
        if not raw:
            k = '未上架'
        elif raw in IT_RACK_MOUNT_STATUS_ENUM:
            k = raw
        else:
            k = _coerce_rack_mount_status(raw) or '未上架'
        bucket[k] += 1
    chart = [{'label': lab, 'count': bucket[lab]} for lab in sorted(bucket.keys(), key=lambda x: (-bucket[x], x))]
    return jsonify({'success': True, 'data': data, 'total': len(data), 'chart': chart})


@app.route('/api/it_cabinets/<int:device_id>', methods=['PUT'])
@login_required
@admin_required
def api_it_cabinet_update(device_id):
    device = Device.query.get_or_404(device_id)
    if not device.category or device.category.name != 'IT机柜':
        return jsonify({'success': False, 'message': '仅支持 IT机柜 台账'}), 400
    payload = request.get_json() or {}
    st = (payload.get('it_rack_mount_status') or '').strip()
    if st not in IT_RACK_MOUNT_STATUS_ENUM:
        return jsonify({'success': False, 'message': f'上下架状态必须是：{", ".join(IT_RACK_MOUNT_STATUS_ENUM)}'}), 400
    device.it_rack_mount_status = st
    device.updated_at = datetime.now()
    db.session.commit()
    return jsonify({'success': True, 'message': '已更新', 'data': serialize_device(device)})


def _category_is_it_cabinet(category_id) -> bool:
    if not category_id:
        return False
    c = DeviceCategory.query.get(category_id)
    return bool(c and (c.name or '').strip() == 'IT机柜')


def _apply_devices_status_filter(query, category_id, status: str):
    """IT机柜 按 it_rack_mount_status 筛选；其它分类按 asset_status。"""
    if not (status or '').strip():
        return query
    status = status.strip()
    if _category_is_it_cabinet(category_id) and status in IT_RACK_MOUNT_STATUS_ENUM:
        if status == '未上架':
            return query.filter(
                or_(
                    Device.it_rack_mount_status == '未上架',
                    Device.it_rack_mount_status.is_(None),
                    Device.it_rack_mount_status == '',
                )
            )
        return query.filter(Device.it_rack_mount_status == status)
    # 资产状态：支持「使用中」作为前缀筛选（包含 使用中-已上电...）
    if status == '使用中':
        return query.filter(Device.asset_status.isnot(None), Device.asset_status.like('使用中%'))
    return query.filter_by(asset_status=status)


# ============== 路由 - 设备管理 ==============

@app.route('/devices')
@login_required
def devices():
    category_id = request.args.get('category', type=int)
    search = request.args.get('search', '')
    status = request.args.get('status', '')
    expired = request.args.get('expired', '')
    
    query = Device.query

    # 专业权限过滤（基于回填后的 Device.profession）
    allowed = allowed_professions_for_current_user()
    if allowed is not None and allowed:
        query = query.filter(Device.profession.in_(list(allowed)))
    
    if category_id:
        query = query.filter_by(category_id=category_id)
    if search:
        query = query.filter(
            (Device.instance_name.contains(search)) |
            (Device.unique_id.contains(search)) |
            (Device.location.contains(search)) |
            (Device.brand.contains(search)) |
            (Device.model.contains(search))
        )
    query = _apply_devices_status_filter(query, category_id, status)
    if str(expired).lower() in ('1', 'true', 'yes', 'y'):
        query = query.filter_by(is_expired=True)
    
    devices_list = query.order_by(Device.updated_at.desc()).limit(500).all()
    categories = DeviceCategory.query.all()
    it_cat = DeviceCategory.query.filter_by(name='IT机柜').first()
    it_cabinet_category_id = it_cat.id if it_cat else None

    return render_template('devices.html', 
                         devices=devices_list, 
                         categories=categories,
                         current_category=category_id,
                         search=search,
                         status=status,
                         expired=expired,
                         it_cabinet_category_id=it_cabinet_category_id)

@app.route('/api/devices')
@login_required
def api_devices():
    category_id = request.args.get('category', type=int)
    search = request.args.get('search', '')
    status = request.args.get('status', '')
    expired = request.args.get('expired', '')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    
    query = Device.query

    allowed = allowed_professions_for_current_user()
    if allowed is not None and allowed:
        query = query.filter(Device.profession.in_(list(allowed)))
    
    if category_id:
        query = query.filter_by(category_id=category_id)
    if search:
        query = query.filter(
            (Device.instance_name.contains(search)) |
            (Device.unique_id.contains(search)) |
            (Device.location.contains(search)) |
            (Device.brand.contains(search)) |
            (Device.model.contains(search))
        )
    query = _apply_devices_status_filter(query, category_id, status)
    if str(expired).lower() in ('1', 'true', 'yes', 'y'):
        query = query.filter_by(is_expired=True)
    
    pagination = query.order_by(Device.updated_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    devices_data = [serialize_device(d) for d in pagination.items]
    
    return jsonify({
        'success': True,
        'data': devices_data,
        'total': pagination.total,
        'pages': pagination.pages,
        'page': page
    })

@app.route('/device/<int:device_id>')
@login_required
def device_detail(device_id):
    device = Device.query.get_or_404(device_id)
    convert_params = lookup_convert_params(
        warehouse_dir=DEVICE_CONVERT_DIR,
        unique_id=device.unique_id,
        model=device.model,
    )
    return render_template(
        'device_detail.html', device=device, convert_params=convert_params
    )

@app.route('/api/device/<int:device_id>')
@login_required
def api_device_detail(device_id):
    device = Device.query.get_or_404(device_id)
    events = AssetEvent.query.filter_by(device_id=device.id).order_by(AssetEvent.occurred_at.desc()).limit(200).all()
    return jsonify({'success': True, 'data': serialize_device(device), 'events': [serialize_event(e) for e in events]})


@app.route('/api/devices/<int:device_id>/convert-params', methods=['GET'])
@login_required
def api_device_convert_params(device_id):
    device = Device.query.get_or_404(device_id)
    data = lookup_convert_params(
        warehouse_dir=DEVICE_CONVERT_DIR,
        unique_id=device.unique_id,
        model=device.model,
    )
    return jsonify({'success': True, 'data': data})

@app.route('/api/devices/<int:device_id>/events', methods=['GET', 'POST'])
@login_required
def api_device_events(device_id):
    device = Device.query.get_or_404(device_id)
    if request.method == 'GET':
        events = AssetEvent.query.filter_by(device_id=device.id).order_by(AssetEvent.occurred_at.desc()).limit(500).all()
        return jsonify({'success': True, 'data': [serialize_event(e) for e in events]})

    if not is_admin():
        return jsonify({'success': False, 'message': '需要管理员权限'}), 403

    data = request.get_json() or {}
    event_type = (data.get('event_type') or 'note').strip()
    title = (data.get('title') or '').strip()
    note = (data.get('note') or '').strip()
    status = (data.get('status') or 'done').strip()
    occurred_at = parse_datetime(data.get('occurred_at')) or datetime.now()
    payload = data.get('payload')

    e = AssetEvent(
        device_id=device.id,
        event_type=event_type,
        title=title or None,
        note=note or None,
        status=status,
        occurred_at=occurred_at,
        created_by=current_user.id if current_user.is_authenticated else None,
        payload_json=json.dumps(payload, ensure_ascii=False) if payload is not None else None,
    )
    db.session.add(e)
    db.session.commit()
    return jsonify({'success': True, 'message': '事件已记录', 'data': serialize_event(e)})

# ============== 路由 - 数据导入导出 ==============

@app.route('/maintenance')
@login_required
@admin_required
def maintenance():
    categories = DeviceCategory.query.all()
    return render_template(
        'maintenance.html',
        categories=categories,
        warehouse_dir_default=WAREHOUSE_DIR,
    )

@app.route('/sync_staff', methods=['POST'])
@login_required
@admin_required
def sync_staff():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '没有上传文件'})
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '没有选择文件'})

    filename = secure_filename(file.filename)
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(file_path)

    try:
        rows = _read_excel_rows(file_path)
        # 写入标准人员名单Excel（覆盖）
        write_staff_excel(rows, STAFF_FILE)

        # 同步到用户表
        stats = sync_users_from_staff_rows(rows)

        return jsonify({
            'success': True,
            'message': f"同步完成：共{stats['total']}行，新增{stats['inserted']}，更新{stats['updated']}，跳过{stats['skipped']}",
            'data': stats
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'同步失败: {str(e)}'})
    finally:
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
        except Exception:
            pass

@app.route('/import', methods=['GET', 'POST'])
@login_required
@admin_required
def import_devices():
    if request.method == 'POST':
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '没有上传文件'})
        
        file = request.files['file']
        category_name = (request.form.get('category') or '').strip()
        mode = (request.form.get('mode') or 'insert').strip().lower()  # insert | upsert
        
        if file.filename == '':
            return jsonify({'success': False, 'message': '没有选择文件'})
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            
            try:
                rows = _read_excel_rows(file_path)
                count = 0
                updated = 0
                inserted = 0

                # 如果文件里包含“设备分类”，优先按每行分类处理；否则使用表单分类
                for row in rows:
                    row_category_name = _to_str(row.get('设备分类')) or category_name or '其他设备'

                    category = DeviceCategory.query.filter_by(name=row_category_name).first()
                    if not category:
                        category = DeviceCategory(
                            name=row_category_name,
                            display_name=row_category_name,
                            description=f'{row_category_name}设备台账'
                        )
                        db.session.add(category)
                        db.session.commit()

                    unique_id = _unique_id_from_row(row)
                    if mode == 'upsert' and unique_id:
                        existing = Device.query.filter_by(category_id=category.id, unique_id=unique_id).first()
                        if existing:
                            update_device_from_row(existing, row)
                            updated += 1
                        else:
                            device = Device(category_id=category.id)
                            update_device_from_row(device, row)
                            db.session.add(device)
                            inserted += 1
                    else:
                        device = Device(category_id=category.id)
                        update_device_from_row(device, row)
                        db.session.add(device)
                        inserted += 1
                    count += 1
                
                db.session.commit()
                
                return jsonify({
                    'success': True, 
                    'message': (
                        f'处理完成：共 {count} 条；新增 {inserted} 条'
                        + (f'；更新 {updated} 条' if mode == 'upsert' else '')
                    ),
                    'category': category_name or None,
                    'mode': mode
                })
                
            except Exception as e:
                db.session.rollback()
                return jsonify({'success': False, 'message': f'导入失败: {str(e)}'})
            finally:
                if os.path.exists(file_path):
                    os.remove(file_path)
        else:
            return jsonify({'success': False, 'message': '不支持的文件格式'})
    
    return redirect(url_for('maintenance'))

@app.route('/reload')
@login_required
@admin_required
def reload_devices():
    """重新从原始Excel文件加载设备"""
    try:
        # 清空现有设备
        Device.query.delete()
        db.session.commit()
        
        # 重新加载
        stats = load_devices_from_warehouse(
            warehouse_dir=WAREHOUSE_DIR,
            update_device_from_row=update_device_from_row,
            merge_it_cabinet_rows=_merge_it_cabinet_pkl_rows_by_unique_id,
            normalize_it_rack_mount_status_for_it_cabinets=normalize_it_rack_mount_status_for_it_cabinets,
            prune_missing=False,
        )
        if stats.get('files', 0) == 0:
            load_devices_from_excel()
        
        return jsonify({'success': True, 'message': '设备数据已重新加载', 'data': stats})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/sync_warehouse', methods=['POST'])
@login_required
@admin_required
def sync_warehouse():
    """
    从仓库源目录批量upsert设备数据。

    可选参数：
    - warehouse_dir: 自定义目录（否则用环境变量/默认目录）
    - prune_missing: true/false，是否删除同分类下源中不存在的记录（谨慎开启）
    """
    data = request.get_json(silent=True) or request.form or {}
    warehouse_dir = (data.get('warehouse_dir') or '').strip() or None
    prune_missing = str(data.get('prune_missing') or '').strip().lower() in ('1', 'true', 'yes', 'y', 'on')

    try:
        stats = load_devices_from_warehouse(
            warehouse_dir=warehouse_dir or WAREHOUSE_DIR,
            update_device_from_row=update_device_from_row,
            merge_it_cabinet_rows=_merge_it_cabinet_pkl_rows_by_unique_id,
            normalize_it_rack_mount_status_for_it_cabinets=normalize_it_rack_mount_status_for_it_cabinets,
            prune_missing=prune_missing,
        )
        return jsonify({'success': True, 'message': '仓库源同步完成', 'data': stats})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'仓库源同步失败: {str(e)}'})


@app.route('/remap_device_database', methods=['POST'])
@login_required
@admin_required
def remap_device_database():
    """
    从 device_database（或其它目录）全量同步 pkl，并规范化 IT机柜「上下架」三枚举。
    默认目录：环境变量 ASSET_WAREHOUSE_DIR 或 项目下 device_database。
    请求 JSON：{ "warehouse_dir": "E:\\\\...\\\\device_database" } 可选。
    """
    data = request.get_json(silent=True) or {}
    warehouse_dir = (data.get('warehouse_dir') or '').strip() or WAREHOUSE_DIR
    prune_missing = str(data.get('prune_missing') or '').strip().lower() in ('1', 'true', 'yes', 'y', 'on')
    try:
        stats = load_devices_from_warehouse(
            warehouse_dir=warehouse_dir,
            update_device_from_row=update_device_from_row,
            merge_it_cabinet_rows=_merge_it_cabinet_pkl_rows_by_unique_id,
            normalize_it_rack_mount_status_for_it_cabinets=normalize_it_rack_mount_status_for_it_cabinets,
            prune_missing=prune_missing,
        )
        msg = (
            f"已自 {stats.get('warehouse_dir')} 同步 {stats.get('files')} 个 pkl；"
            f"IT机柜上下架规范化 {stats.get('it_rack_mount_normalized', 0)} 条"
        )
        return jsonify({'success': True, 'message': msg, 'data': stats})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/export')
@login_required
def export_devices():
    """导出设备数据（支持按分类/位置信息筛选）"""
    if not can_export():
        return jsonify({'success': False, 'message': '无导出权限'}), 403

    category_id = request.args.get('category', type=int)
    location = (request.args.get('location') or '').strip()
    location_mode = (request.args.get('location_mode') or 'contains').strip().lower()  # contains|prefix|exact
    
    query = Device.query
    if category_id:
        query = query.filter_by(category_id=category_id)
    if location:
        if location_mode == 'exact':
            query = query.filter(Device.location == location)
        elif location_mode == 'prefix':
            query = query.filter(Device.location.isnot(None), Device.location.like(f"{location}%"))
        else:
            # default: contains
            query = query.filter(Device.location.isnot(None), Device.location.contains(location))
    
    devices = query.all()

    headers = [
        '范围域号', '唯一ID', '实例名称', '资产状态', 'IT机柜上下架状态', '位置信息', '品牌', '型号', '质保年限', '生命周期',
        '投产日期', '质保开始时间', '质保结束时间', '是否过保', '能耗类型', '阈值范围', '动环设备类型',
        '额定电流容量', '二维码', '设备分类'
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = 'devices'
    ws.append(headers)
    for d in devices:
        ws.append([
            d.scope_id or '',
            d.unique_id or '',
            d.instance_name or '',
            d.asset_status or '',
            d.it_rack_mount_status or '',
            d.location or '',
            d.brand or '',
            d.model or '',
            d.warranty_years or '',
            d.lifecycle or '',
            d.commission_date.isoformat() if d.commission_date else '',
            d.warranty_start.isoformat() if d.warranty_start else '',
            d.warranty_end.isoformat() if d.warranty_end else '',
            '是' if d.is_expired else '否',
            d.energy_type or '',
            d.threshold_range or '',
            d.device_type or '',
            d.rated_capacity or '',
            d.qr_code or '',
            d.category.name if d.category else ''
        ])

    export_path = os.path.join(app.config['UPLOAD_FOLDER'], 'devices_export.xlsx')
    wb.save(export_path)
    # 下载名：包含筛选条件（避免用户混淆不同导出）
    name_parts = ['devices']
    if category_id:
        name_parts.append(f"cat{category_id}")
    if location:
        safe_loc = re.sub(r'[^0-9A-Za-z\u4e00-\u9fff_-]+', '_', location)[:40]
        name_parts.append(f"loc_{safe_loc}")
        name_parts.append(location_mode)
    download_name = "_".join(name_parts) + ".xlsx"
    return send_file(export_path, as_attachment=True, download_name=download_name)

# ============== 路由 - 用户管理 ==============

@app.route('/users')
@login_required
@admin_required
def users():
    users_list = User.query.all()
    return render_template('users.html', users=users_list)

@app.route('/api/users', methods=['GET', 'POST'])
@login_required
@admin_required
def api_users():
    if request.method == 'GET':
        users_list = User.query.all()
        return jsonify({
            'success': True,
            'data': [{
                'id': u.id,
                'username': u.username,
                'display_name': u.display_name,
                'employee_id': u.employee_id,
                'role': u.role,
                'is_active': u.is_active,
                'created_at': u.created_at.isoformat()
            } for u in users_list]
        })
    
    # POST - 创建/更新用户
    data = request.get_json()
    
    if 'id' in data and data['id']:
        # 更新
        user = User.query.get(data['id'])
        if not user:
            return jsonify({'success': False, 'message': '用户不存在'})
        
        if 'display_name' in data:
            user.display_name = data['display_name']
        if 'role' in data:
            user.role = data['role']
        if 'is_active' in data:
            user.is_active = data['is_active']
        if 'password' in data and data['password']:
            user.set_password(data['password'])
    else:
        # 创建
        if not data.get('username') or not data.get('password'):
            return jsonify({'success': False, 'message': '用户名和密码不能为空'})
        
        if User.query.filter_by(username=data['username']).first():
            return jsonify({'success': False, 'message': '用户名已存在'})
        
        user = User(
            username=data['username'],
            display_name=data.get('display_name', data['username']),
            employee_id=data.get('employee_id'),
            role=data.get('role', 'user')
        )
        user.set_password(data['password'])
        db.session.add(user)
    
    db.session.commit()
    return jsonify({'success': True, 'message': '保存成功'})

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    
    if user.username == 'admin':
        return jsonify({'success': False, 'message': '不能删除管理员账户'})
    
    db.session.delete(user)
    db.session.commit()
    return jsonify({'success': True, 'message': '用户已删除'})

# ============== 路由 - 设备维护 ==============

@app.route('/api/devices/<int:device_id>', methods=['PUT', 'DELETE'])
@login_required
@admin_required
def manage_device(device_id):
    device = Device.query.get_or_404(device_id)
    
    if request.method == 'DELETE':
        db.session.delete(device)
        db.session.commit()
        return jsonify({'success': True, 'message': '设备已删除'})
    
    # PUT - 更新
    data = request.get_json()
    
    if 'instance_name' in data:
        device.instance_name = data['instance_name']
    if 'asset_status' in data:
        new_status = (data.get('asset_status') or '').strip()
        if new_status and new_status not in ASSET_STATUS_ENUM:
            return jsonify({'success': False, 'message': f'资产状态不合法：{new_status}'}), 400
        device.asset_status = new_status or None
    if 'location' in data:
        device.location = data['location']
    if 'brand' in data:
        device.brand = data['brand']
    if 'model' in data:
        device.model = data['model']
    if 'warranty_years' in data:
        device.warranty_years = data['warranty_years']
    if 'commission_date' in data:
        device.commission_date = parse_date(data['commission_date'])
    if 'warranty_start' in data:
        device.warranty_start = parse_date(data['warranty_start'])
    if 'warranty_end' in data:
        device.warranty_end = parse_date(data['warranty_end'])
    if 'energy_type' in data:
        device.energy_type = data['energy_type']
    if 'device_type' in data:
        device.device_type = data['device_type']
    if 'rated_capacity' in data:
        device.rated_capacity = data['rated_capacity']
    
    if device.warranty_end:
        device.is_expired = datetime.now().date() > device.warranty_end
    
    device.updated_at = datetime.now()
    db.session.commit()
    
    return jsonify({'success': True, 'message': '设备已更新'})

@app.route('/api/categories', methods=['GET', 'POST'])
@login_required
@admin_required
def api_categories():
    if request.method == 'GET':
        categories = DeviceCategory.query.all()
        return jsonify({
            'success': True,
            'data': [{
                'id': c.id,
                'name': c.name,
                'display_name': c.display_name,
                'description': c.description,
                'device_count': Device.query.filter_by(category_id=c.id).count()
            } for c in categories]
        })
    
    data = request.get_json()
    
    if 'id' in data and data['id']:
        category = DeviceCategory.query.get(data['id'])
        if category:
            if 'display_name' in data:
                category.display_name = data['display_name']
            if 'description' in data:
                category.description = data['description']
    else:
        category = DeviceCategory(
            name=data.get('name', data.get('display_name')),
            display_name=data.get('display_name'),
            description=data.get('description', '')
        )
        db.session.add(category)
    
    db.session.commit()
    return jsonify({'success': True, 'message': '保存成功'})


@app.route('/api/admin/reimport_power_pkls', methods=['POST'])
@login_required
@admin_required
def api_reimport_power_pkls():
    """
    一键清空并用“上电 pkl”重灌：
    - 交流PDU：E:\\DCIM_AI\\device_monitor\\device_database\\R楼-PDU上电.pkl（唯一ID末尾 1.1.7.4）
    - IT机柜：E:\\DCIM_AI\\device_monitor\\device_database\\R楼-IT机柜上电.pkl（唯一ID末尾 2.7.1.1）
    """
    data = request.get_json(silent=True) or {}
    pdu_pkl_path = (data.get("pdu_pkl_path") or r"E:\DCIM_AI\device_monitor\device_database\R楼-PDU上电.pkl").strip()
    it_pkl_path = (data.get("it_cabinet_pkl_path") or r"E:\DCIM_AI\device_monitor\device_database\R楼-IT机柜上电.pkl").strip()
    try:
        res = reimport_power_pkls(
            pdu_pkl_path=pdu_pkl_path,
            it_cabinet_pkl_path=it_pkl_path,
            update_device_from_row=update_device_from_row,
            pdu_uid_suffix="1.1.7.4",
            it_uid_suffix="2.7.1.1",
        )
        return jsonify({"success": True, "message": "已清空并重新导入", "data": res})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500

# ============== 初始化 ==============

def init_db():
    """初始化数据库"""
    with app.app_context():
        db.create_all()
        ensure_device_schema_sqlite()
        init_users()
        # 优先从“仓库源目录”同步；若目录不存在/为空，回退到固定文件清单
        stats = load_devices_from_warehouse(
            warehouse_dir=WAREHOUSE_DIR,
            update_device_from_row=update_device_from_row,
            merge_it_cabinet_rows=_merge_it_cabinet_pkl_rows_by_unique_id,
            normalize_it_rack_mount_status_for_it_cabinets=normalize_it_rack_mount_status_for_it_cabinets,
            prune_missing=False,
        )
        if stats.get('files', 0) == 0:
            load_devices_from_excel()
            apply_post_warehouse_category_fixes()
        maybe_hydrate_it_cabinets_from_pkl()
        normalize_it_rack_mount_status_for_it_cabinets()
        print("数据库初始化完成!")

if __name__ == '__main__':
    skip_init = str(os.environ.get('SKIP_INIT_DB', '')).strip().lower() in ('1', 'true', 'yes', 'y', 'on')
    if not skip_init:
        init_db()
    host = os.environ.get('HOST', '0.0.0.0')
    try:
        port = int(os.environ.get('PORT', '5001'))
    except Exception:
        port = 5001
    app.run(debug=True, host=host, port=port)
