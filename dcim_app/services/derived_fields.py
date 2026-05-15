from __future__ import annotations

import os
import unicodedata

from openpyxl import load_workbook
from sqlalchemy import or_

from ..extensions import db
from ..models import Device
from ..utils.ids import to_str
from ..utils.schema import ensure_device_schema_sqlite


def parse_unique_id_meta(unique_id: str | None) -> dict:
    """
    规则：
    - 从右往左数第4位：1=设备，2=空间
    - 若为设备：倒数第3位为专业：1电气，2暖通，3弱电，4消防
    """
    uid = to_str(unique_id)
    if not uid:
        return {"object_kind": "unknown", "profession": "未知", "profession_code": None}
    parts = [p for p in str(uid).strip().split(".") if p != ""]
    if len(parts) < 4:
        return {"object_kind": "unknown", "profession": "未知", "profession_code": None}
    try:
        obj_flag = int(parts[-4])
    except Exception:
        obj_flag = None
    object_kind = "unknown"
    if obj_flag == 1:
        object_kind = "device"
    elif obj_flag == 2:
        object_kind = "space"

    profession_code = None
    profession = "未知"
    if object_kind == "device" and len(parts) >= 3:
        try:
            profession_code = int(parts[-3])
        except Exception:
            profession_code = None
        profession = {1: "电气", 2: "暖通", 3: "弱电", 4: "消防"}.get(
            profession_code, "未知"
        )

    return {
        "object_kind": object_kind,
        "profession": profession,
        "profession_code": profession_code,
    }


_MODEL_CLASSIFICATION_CACHE = {"loaded": False, "map": {}}


def load_model_classification_map() -> dict[str, str]:
    """
    从 xlsx 读取“型号/模型 -> 分类”映射。
    默认路径：
    - 环境变量 MODEL_CLASSIFICATION_XLSX
    - E:\\世纪互联\\太仓\\设备管理AI\\data_center_facility_manager\\设备模型分类.xlsx
    """
    if _MODEL_CLASSIFICATION_CACHE["loaded"]:
        return _MODEL_CLASSIFICATION_CACHE["map"]
    _MODEL_CLASSIFICATION_CACHE["loaded"] = True

    path = os.environ.get("MODEL_CLASSIFICATION_XLSX") or r"E:\世纪互联\太仓\设备管理AI\data_center_facility_manager\设备模型分类.xlsx"
    if not path or not os.path.exists(path):
        _MODEL_CLASSIFICATION_CACHE["map"] = {}
        return {}
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, max_row=2000, values_only=True))
        if not rows:
            _MODEL_CLASSIFICATION_CACHE["map"] = {}
            return {}
        headers = [
            unicodedata.normalize("NFKC", str(h).strip()) if h is not None else ""
            for h in rows[0]
        ]
        key_idx = None
        val_idx = None
        for i, h in enumerate(headers):
            if not h:
                continue
            if key_idx is None and (("型号" in h) or ("模型" in h)):
                key_idx = i
            if val_idx is None and ("分类" in h):
                val_idx = i
        if key_idx is None or val_idx is None:
            _MODEL_CLASSIFICATION_CACHE["map"] = {}
            return {}
        mp: dict[str, str] = {}
        for r in rows[1:]:
            if not r:
                continue
            k = to_str(r[key_idx] if key_idx < len(r) else None)
            v = to_str(r[val_idx] if val_idx < len(r) else None)
            if not k or not v:
                continue
            mp[k] = v
        _MODEL_CLASSIFICATION_CACHE["map"] = mp
        return mp
    except Exception:
        _MODEL_CLASSIFICATION_CACHE["map"] = {}
        return {}


def model_to_classification(model: str | None) -> str | None:
    m = to_str(model)
    if not m:
        return None
    mp = load_model_classification_map()
    if not mp:
        return None
    return mp.get(m)


def backfill_uid_derived_fields(limit: int = 2000) -> int:
    """
    为旧数据回填 object_kind/profession/model_classification。
    """
    try:
        ensure_device_schema_sqlite()
    except Exception:
        pass

    q = Device.query.filter(
        Device.unique_id.isnot(None),
        Device.unique_id != "",
        or_(
            Device.object_kind.is_(None),
            Device.object_kind == "",
            Device.profession.is_(None),
            Device.profession == "",
        ),
    ).limit(limit)
    rows = q.all()
    changed = 0
    for d in rows:
        meta = parse_unique_id_meta(d.unique_id)
        if not d.object_kind:
            d.object_kind = meta["object_kind"]
        if not d.profession:
            d.profession = meta["profession"]
        if not d.model_classification:
            d.model_classification = model_to_classification(d.model)
        changed += 1
    if changed:
        db.session.commit()
    return changed

