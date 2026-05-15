from __future__ import annotations

import pickle
import unicodedata

from openpyxl import load_workbook


def read_excel_rows(file_path: str):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    headers = []
    for cell in ws[1]:
        if cell.value is None:
            headers.append("")
        else:
            headers.append(unicodedata.normalize("NFKC", str(cell.value).strip()))
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in r):
            continue
        row = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            row[h] = r[i] if i < len(r) else None
        rows.append(row)
    return rows


def read_pkl_rows(file_path: str):
    with open(file_path, "rb") as f:
        obj = pickle.load(f)

    recs: list = []
    if hasattr(obj, "to_dict"):
        try:
            recs = obj.to_dict(orient="records")
        except TypeError:
            recs = []
    if not recs and isinstance(obj, list):
        recs = [r for r in obj if isinstance(r, dict)]

    if not recs:
        raise ValueError(f"Unsupported pkl content type: {type(obj)}")

    out = []
    for r in recs:
        if not isinstance(r, dict):
            continue
        nr = {}
        for k, v in r.items():
            if k is None:
                continue
            nk = unicodedata.normalize("NFKC", str(k).strip())
            if nk:
                nr[nk] = v
        out.append(nr)
    return out

